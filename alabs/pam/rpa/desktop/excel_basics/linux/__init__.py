#!/usr/bin/env python
# coding=utf8
"""
====================================
 :mod:`alabs.pam.la.bot2py
====================================
.. moduleauthor:: Raven Lim <deokyu@argos-labs.com>
.. note:: VIVANS License

Description
===========
ARGOS LABS PAM For LA

Authors
===========

* Raven Lim

Change Log
--------

 * [2019/01/30]
    - starting
"""

################################################################################
import sys
import re
import traceback

import openpyxl
from openpyxl.worksheet import worksheet
from alabs.common.util.vvargs import ModuleContext, func_log
from alabs.common.util.vvlogger import StructureLogFormat



################################################################################
# Version
NUM_VERSION = (0, 9, 0)
VERSION = ".".join(str(nv) for nv in NUM_VERSION)
__version__ = VERSION

OWNER = 'ARGOS-LABS'
GROUP = 'Pam'
PLATFORM = ['windows', 'darwin', 'linux']
OUTPUT_TYPE = 'json'
DESCRIPTION = 'Pam for HA. It reads json scenario files by LA Stu and runs'

import openpyxl
import json
import ntpath

################################################################################
def excel_column_calculate(n:int, d:list):
    """
    Excel의 `A`, `AA` 와 같은 값을 Column 값을 쉽게 구할 수 있습니다.
    :param n: 첫 번째 Column 기준으로 1 이상의 수
    :param d: 빈 list
    :return: str
    >>> excel_column_calculate(1, []))
    'A'
    >>> excel_column_calculate(27, [])
    'AA'
    >>> excel_column_calculate(1000, [])
    'ALL'
    """
    if not n:
        return ''.join(d)
    n -= 1
    r = n // 26
    m = (n % 26)
    d.insert(0, chr(65 + m))
    return excel_column_calculate(r, d)


################################################################################
def excel_column_calculator_str_to_num(col: str):
    """
    >>> col = 'A'
    >>> excel_column_calculator_str_to_num(col)
    1
    >>> col = 'AA'
    >>> excel_column_calculator_str_to_num(col)
    27
    >>> col = 'ALL'
    >>> excel_column_calculator_str_to_num(col)
    1000
    """
    cols = list(col)
    data = list()
    cols.reverse()
    for i, c in list(enumerate(cols)):
        data.append((26 ** i) * (ord(c.upper()) - 64))
    return sum(data)


################################################################################
def _open_excel_file(filename) -> openpyxl.Workbook:
    """
    파라메터 filename의 위치에 있는 엑셀파일을 열어서
    openpyxl.Workbook 객체를 반환
    :param filename: .xlsx 형태의 엑셀파일
    :return: <openpyxl.Workbook>
    """
    return openpyxl.load_workbook(filename)


################################################################################
def _read_sheet(wb:openpyxl.Workbook, sheet:str):
    """
    파라메터 sheet를 Workbook에서 찾아서 반환
    :param wb: <openpyxl.Workbook>
    :param sheet: WorkSheet 이름
    :return: <openpyxl.Worksheet>
    """
    return wb[sheet]

################################################################################
def _read_head(ws:openpyxl.worksheet.worksheet, data_range):
    # 헤드 범위가 지정되지 않았을 경우
    # 첫 번째 줄의 시작과 끝의 위치를 가지고 시작
    if not data_range:
        max_column = excel_column_calculate(ws.max_column, [])
        data_range = "A1:{}1".format(max_column)
    return [x.value for x in ws[data_range] for x in x]

################################################################################
def _read_data(ws:openpyxl.worksheet.worksheet, data_range=None):
    """

    :param ws:
    :param data_range:
    :return:
        A1:B2
            [['ABC', 'DEF'], ['AAA', 'BBB']]
    """
    data = []
    if not data_range:
        max_column = excel_column_calculate(ws.max_column, [])
        data_range = "A1:{}{}".format(max_column, ws.max_row)

    cells = ws[data_range]
    # if not isinstance(cells, ):
    try:
        for row in cells:
            value = []
            for cell in row:
                value.append(cell.value)
            data.append(value)
    except TypeError:
        data = [[cells.value]]
    return data

################################################################################
def _to_dict(head: list, data: list):
    r = []
    for d in data:
        r.append(dict(zip(head, d)))
    return r

################################################################################
def _to_json(data: list):
    return json.dumps(data, indent=4, ensure_ascii=False)

################################################################################
def _as_file(path:str, data):
    head, tail = ntpath.split(path)
    filename_ext = tail or ntpath.basename(head)
    filename, ext = ntpath.splitext(filename_ext)
    with open(filename + '.json', 'w') as f:
        f.write(data)

def seperate_str_num(taget):
    """
    >>> taget = 'A1'
    >>> print(seperate_str_num('A1'))
    ('A', '1')
    """
    match = re.match(r"([a-z]+)([0-9]+)", taget, re.I)
    if match:
        return match.groups()
    raise ValueError(f'{taget} is not a correct pointer for excel. ex) A1')


################################################################################
def read(ws, request):
    result = list()
    for d in request:
        data = json.loads(d)
        data = _read_data(ws, data['cell'])
        data = json.dumps(data, ensure_ascii=False)
        result.append(json.loads(data))
        # result.append(value)
    return result


################################################################################
def write(ws, wb,  filename, request):
    for d in request:
        data = json.loads(d)
        cell = data['cell']
        r, c = ws[cell].row, ws[cell].column

        ori = data['orientation']
        if ori == 'vertical':
            s = r
        else:
            s = c
        for i, v in enumerate(data['values'], s):
            if ori == 'vertical':
                ws.cell(i, c).value = v
            else:
                ws.cell(r, i).value = v
    wb.save(filename)


################################################################################
@func_log
def excel_basics(mcxt, argspec):
    """
    LocateImage
    :param mcxt: module context
    :param argspec: argument spec
    :return: x, y
    """
    mcxt.logger.info("Excel Basic start ...")

    status = True
    message = ''
    value = None

    try:
        wb = _open_excel_file(argspec.filename)
        ws = _read_sheet(wb, argspec.worksheet)

        if argspec.read:
            value = read(ws, argspec.read)
        else:
            write(ws, wb, argspec.filename, argspec.write)
        wb.close()

    except Exception as e:
        status = False
        print(traceback.print_exc())
        mcxt.logger.exception(str(e))
        message = str(e)

    result = StructureLogFormat(RETURN_CODE=status,
                                RETURN_VALUE={
                                    'RESULT': status, 'VALUE': value},
                                MESSAGE=message)
    std = {True: sys.stdout, False: sys.stderr}[status]
    std.write(str(result))

    mcxt.logger.info("Send Email end ...")
    return result


################################################################################
def _main(*args):
    """
    Build user argument and options and call plugin job function
    :param args: user arguments
    :return: return value from plugin job function

    ..note:: _main 함수에서 사용되는 패러미터(옵션) 정의 방법
플러그인 모듈은 ModuleContext 을 생성하여 mcxt를 with 문과 함께 사용
    owner='ARGOS-LABS',
    group='pam',
    version='1.0',
    platform=['windows', 'darwin', 'linux'],
    output_type='text',
    description='HA Bot for LA',
    test_class=TU,
    """
    with ModuleContext(
        owner=OWNER,
        group=GROUP,
        version=VERSION,
        platform=PLATFORM,
        output_type=OUTPUT_TYPE,
        description=DESCRIPTION,
    ) as mcxt:
        # python -m alabs.pam.rpa.desktop.excel_basics
        # filename worksheet
        # --read {"cell": "A1:B2"}
        # --read {"cell": "B2"}
        # --write {"cell": "B3",  "orientation": "vertical", "values": ["value", "value", "value"]}
        # --write {"cell": "B3",  "orientation": "vertical", "values": ["value", "value", "value"]}


        # 필수 입력 항목
        mcxt.add_argument('filename', type=str, help='')
        mcxt.add_argument('worksheet', type=str, help='')
        mcxt.add_argument('--write', type=str, action='append', help='', default=list())
        mcxt.add_argument('--read', type=str, action='append', help='', default=list())

        argspec = mcxt.parse_args(args)
        return excel_basics(mcxt, argspec)


################################################################################
def main(*args):
    return _main(*args)

